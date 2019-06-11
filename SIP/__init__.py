import os
import time
import bagit
import shutil
import shortuuid
from datetime import datetime

class SubmissionInformationPackage:

    def __init__(self):
        pass
        
    def load(self, path):
        if not os.path.isdir(path):
            raise Exception("ERROR: " + str(path) + " is not a valid SIP. You may want to create a SIP with .create().")

        self.bag = bagit.Bag(path)
        self.bagID = os.path.basename(path)
        self.colID = self.bagID.split("_")[0]
        self.data = os.path.join(path, "data")
        
    
    def create(self, colID):
        sipPath= "/media/Masters/Archives/SIP"
        
        metadata = {\
        'Bag-Type': 'SIP', \
        'Bagging-Date': str(datetime.now().isoformat()), \
        'Posix-Date': str(time.time()), \
        'BagIt-Profile-Identifier': 'https://archives.albany.edu/static/bagitprofiles/sip-profile-v0.2.json', \
        'Collection-Identifier': colID \
        }
        
        self.colID = colID
        self.bagID = colID + "_" + str(shortuuid.uuid())
        metadata["Bag-Identifier"] = self.bagID
        if not os.path.isdir(os.path.join(sipPath, colID)):
            os.mkdir(os.path.join(sipPath, colID))
            
        self.bagDir = os.path.join(sipPath, colID, self.bagID)
        os.mkdir(self.bagDir)

        self.bag = bagit.make_bag(self.bagDir, metadata)
        self.data = os.path.join(self.bagDir, "data")
        
        
    def package(self, dir):
        self.setupProcecssing()
        
        excludeList = ["thumbs.db", "desktop.ini", ".ds_store"]
        if not os.path.isdir(dir):
            raise Exception("ERROR: " + str(dir) + " is not a valid path.")
        else:
            # Move files and folders to AIP
            for thing in os.listdir(dir):
                thingPath = os.path.join(dir, thing)
                if os.path.isfile(thingPath):
                    if not thing.lower() in excludeList:
                        shutil.copy2(thingPath, self.data)
                        shutil.copy2(thingPath, self.procMasters)
                else:
                    shutil.copytree(thingPath, os.path.join(self.data, thing))
                    shutil.copytree(thingPath, os.path.join(self.procMasters, thing))
            # Delete files and fodlers after move 
            for thing in os.listdir(dir):
                thingPath = os.path.join(dir, thing)
                if os.path.isfile(thingPath):
                    os.remove(thingPath)
                else:
                    shutil.rmtree(thingPath)
            if len(os.listdir(dir)) == 0:
                os.rmdir(dir)
                
                
    def inventory(self):
        inventory = []
        for root, dirs, files in os.walk(self.data):
            for file in files:
                filePath = os.path.join(root, file).split(os.path.sep)
                relPath = []
                check = False
                for item in filePath:
                    if check == True:
                        relPath.append(item)
                    if item == "data":
                        check = True
                inventory.append(os.path.sep.join(relPath))
        return "\n".join(inventory)
        
    def manifest(self):
        lines = []
        manifest = os.path.join(self.bagDir, "manifest-sha256.txt")
        with open(manifest, "r") as f:
            for line in f.readlines():
                lines.append(line)
            f.close()
        return lines
        
    def size(self):
        suffixes = ['B', 'KB', 'MB', 'GB', 'TB', 'PB']
        dirSize = 0
        fileCount = 0
        for root, dirs, files in os.walk(self.data):
            for file in files:
                fileCount += 1
                dirSize += os.path.getsize(os.path.join(root, file))
        i = 0
        while dirSize >= 1024 and i < len(suffixes)-1:
            dirSize /= 1024.
            i += 1
        f = ('%.2f' % dirSize).rstrip('0').rstrip('.')
        return [f, suffixes[i], fileCount]
                    
    def setupProcecssing(self):
        
        processingRoot = "/media/SPE/processing"
        if not os.path.isdir(processingRoot):
            raise Exception("ERROR: Processing Path " + str(processingRoot) + " is not a valid path.")
        else:
            procCol = os.path.join(processingRoot, str(self.colID))
            if not os.path.isdir(procCol):
                os.mkdir(procCol)
            procPath = os.path.join(procCol, self.bagID)
            if not os.path.isdir(procPath):
                os.mkdir(procPath)
            procMasters = os.path.join(procPath, "masters")
            if not os.path.isdir(procMasters):
                os.mkdir(procMasters)
            self.procMasters = procMasters
            procDerivatives = os.path.join(procPath, "derivatives")
            if not os.path.isdir(procDerivatives):
                os.mkdir(procDerivatives)
            procMetadata = os.path.join(procPath, "metadata")
            if not os.path.isdir(procMetadata):
                os.mkdir(procMetadata)
                
    def extentLog(self, logFile):
        import openpyxl
        if os.path.isfile(logFile):
            wb = openpyxl.load_workbook(filename=logFile, read_only=False)
            sheet = wb.active
            startRow = int(sheet.max_row) + 1
        else:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet["A1"] = "Date"
            sheet["B1"] = "Collection ID"
            sheet["C1"] = "Type"
            sheet["D1"] = "Package"
            sheet["E1"] = "Files"
            sheet["F1"] = "Extent"
            sheet["G1"] = "Extent Bytes"
            startRow = 2
            
        packageSize = self.size()
        sheet["A" + str(startRow)] = self.bag.info["Bagging-Date"]
        sheet["B" + str(startRow)] = self.bag.info["Collection-Identifier"]
        sheet["C" + str(startRow)] = self.bag.info["Bag-Type"]
        sheet["D" + str(startRow)] = self.bag.info["Bag-Identifier"]
        sheet["E" + str(startRow)] = packageSize[2]
        sheet["F" + str(startRow)] = str(packageSize[0]) + " " + str(packageSize[1])
        sheet["G" + str(startRow)] = self.bag.info["Payload-Oxum"]
            
        wb.save(filename=logFile)
